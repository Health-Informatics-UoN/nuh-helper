import csv
import logging
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

logger = logging.getLogger(__name__)

SCAN_REPORT_FILE_NAME = "ScanReport.xlsx"

FIELD_OVERVIEW_HEADERS = [
    "Table",
    "Field",
    "Description",
    "Type",
    "Max length",
    "N rows",
]

TABLE_OVERVIEW_HEADERS = [
    "Table",
    "Description",
    "N rows",
    "N rows checked",
    "N fields",
    "N fields empty",
]

# Mirrors WhiteRabbit's FieldInfo/StringUtilities type detection so scan
# reports classify columns the same way (EMPTY > DATE > INT > REAL > TEXT/VARCHAR).
_DATE_SEPARATORS = ("/", ".", "-", " ")
_INTEGER_PATTERN = re.compile(r"^[+-]?\d+$")
_REAL_PATTERN = re.compile(r"^[+-]?(\d+\.\d*|\.\d+|\d+)([eE][+-]?\d+)?$")
_MIN_AVERAGE_LENGTH_FOR_TEXT = 100


def _generate_date_formats() -> list[str]:
    formats = []
    for sep in _DATE_SEPARATORS:
        for year in ("%y", "%Y"):
            formats.append(f"{year}{sep}%m{sep}%d")  # YMD
            formats.append(f"%d{sep}%m{sep}{year}")  # DMY
            formats.append(f"%m{sep}%d{sep}{year}")  # MDY
    return formats


_DATE_FORMATS = _generate_date_formats()


def _is_integer(value: str) -> bool:
    return bool(_INTEGER_PATTERN.match(value))


def _is_real(value: str) -> bool:
    return bool(_REAL_PATTERN.match(value))


def _is_date(value: str) -> bool:
    for fmt in _DATE_FORMATS:
        try:
            datetime.strptime(value, fmt)
            return True
        except ValueError:
            continue
    return False


class FieldTypeInfo:
    """Tracks the inferred type and max length of a field as values are scanned."""

    def __init__(self) -> None:
        self.n_processed = 0
        self.empty_count = 0
        self.sum_length = 0
        self.max_length = 0
        self.is_integer = True
        self.is_real = True
        self.is_date = True

    def add(self, value: str) -> None:
        self.n_processed += 1
        self.sum_length += len(value)
        self.max_length = max(self.max_length, len(value))

        trimmed = value.strip()
        if not trimmed:
            self.empty_count += 1
            return

        if self.is_real and not _is_real(trimmed):
            self.is_real = False
        if self.is_integer and not _is_integer(trimmed):
            self.is_integer = False
        if self.is_date and not _is_date(trimmed):
            self.is_date = False

    @property
    def type(self) -> str:
        if self.n_processed == self.empty_count:
            return "EMPTY"
        if self.is_date:
            return "DATE"
        if self.is_integer:
            return "INT"
        if self.is_real:
            return "REAL"

        non_empty = self.n_processed - self.empty_count
        if non_empty and (self.sum_length / non_empty) >= _MIN_AVERAGE_LENGTH_FOR_TEXT:
            return "TEXT"
        return "VARCHAR"


def index_table_names(table_names: list[str]) -> dict[str, str]:
    indexed = {}
    counts = defaultdict(int)

    for name in table_names:
        indexed[name] = name if counts[name] == 0 else f"{name}_{counts[name]}"
        counts[name] += 1

    return indexed


def read_csv_header(csv_path: str) -> list[str]:
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        return next(reader)


def scan_csv_values(
    csv_path: str, min_cell_count: int
) -> tuple[dict[str, list[tuple[str, int]]], int, dict[str, FieldTypeInfo]]:
    counters = defaultdict(Counter)
    type_info = defaultdict(FieldTypeInfo)
    row_count = 0

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row_count += 1
            for field, value in row.items():
                value = value if value is not None else ""
                counters[field][value] += 1
                type_info[field].add(value)

    filtered = {}
    for field, counter in counters.items():
        filtered[field] = [
            (val, cnt) for val, cnt in counter.most_common() if cnt >= min_cell_count
        ]

    return filtered, row_count, dict(type_info)


def generate_scan_report(
    csv_files: list[str],
    output_path: str = SCAN_REPORT_FILE_NAME,
    min_cell_count: int = 1,
) -> str:
    logger.info("Generating scan report for %d table(s)", len(csv_files))

    tables = []

    for csv_file in csv_files:
        csv_file = Path(csv_file)
        header = read_csv_header(csv_file.as_posix())
        logger.info("Scanning '%s' (%d field(s))", csv_file.name, len(header))
        tables.append(
            {"name": csv_file.name, "path": csv_file.as_posix(), "fields": header}
        )

    tables.sort(key=lambda t: t["name"])
    indexed_names = index_table_names([t["name"] for t in tables])

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    table_value_data = {}
    table_type_info = {}
    table_row_counts = {}

    for table in tables:
        table_name_indexed = indexed_names[table["name"]]
        value_data, row_count, type_info = scan_csv_values(
            table["path"], min_cell_count
        )
        table_value_data[table_name_indexed] = value_data
        table_type_info[table_name_indexed] = type_info
        table_row_counts[table_name_indexed] = row_count

    # FIELD_OVERVIEW
    field_sheet = wb.create_sheet("Field Overview")
    field_sheet.append(FIELD_OVERVIEW_HEADERS)

    for table in tables:
        table_name_indexed = indexed_names[table["name"]]
        type_info = table_type_info[table_name_indexed]
        for field in table["fields"]:
            info = type_info.get(field)
            field_type = info.type if info else "EMPTY"
            max_length = info.max_length if info else ""
            field_sheet.append(
                [table_name_indexed, field, "", field_type, max_length, ""]
            )
        field_sheet.append([""])

    # TABLE_OVERVIEW
    table_sheet = wb.create_sheet("Table Overview")
    table_sheet.append(TABLE_OVERVIEW_HEADERS)

    for table in tables:
        table_name_indexed = indexed_names[table["name"]]
        row_count = table_row_counts[table_name_indexed]

        table_sheet.append(
            [table_name_indexed, "", row_count, row_count, len(table["fields"]), -1]
        )

    # VALUE SHEETS
    for table in tables:
        table_name_indexed = indexed_names[table["name"]]
        value_sheet = wb.create_sheet(table_name_indexed)

        fields = table["fields"]
        value_data = table_value_data[table_name_indexed]

        header = []
        for field in fields:
            header.append(field)
            header.append("Frequency")
        value_sheet.append(header)

        max_len = max((len(value_data.get(field, [])) for field in fields), default=0)

        for i in range(max_len):
            row = []
            for field in fields:
                values = value_data.get(field, [])
                if i < len(values):
                    row.append(values[i][0])
                    row.append(values[i][1])
                else:
                    row.append("")
                    row.append("")
            value_sheet.append(row)

    # META
    meta_sheet = wb.create_sheet("_")
    meta_sheet.append(["Key", "Value"])
    meta_sheet.append(["Version", "python-whiterabbit-lite"])
    meta_sheet.append(["Scan started at", datetime.now().isoformat()])
    meta_sheet.append(["Scan finished at", datetime.now().isoformat()])
    meta_sheet.append(["sourceType", "CSV_FILES"])
    meta_sheet.append(["scanValues", True])
    meta_sheet.append(["minCellCount", min_cell_count])

    wb.save(output_path)
    logger.info("Scan report written to '%s'", output_path)
    return output_path
