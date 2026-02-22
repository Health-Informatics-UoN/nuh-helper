"""Excel sheet structure helpers for date-shift (merged cells, read/write with description rows)."""

import contextlib
from typing import Any, cast

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


def _get_row_values_resolving_merged(
    sheet: Worksheet, row_1based: int, max_col: int
) -> list[Any]:
    """
    Return values for one row, resolving merged cells to the top-left cell value.

    openpyxl stores the value only in the first cell of a merged range;
    other cells are MergedCell and have no value. This walks the row and
    fills in the value from the merge range's top-left for each column.
    """
    result: list[Any] = []
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=row_1based, column=col)
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left = sheet.cell(
                        row=merged_range.min_row, column=merged_range.min_col
                    )
                    result.append(top_left.value)
                    break
            else:
                result.append(None)
        else:
            result.append(cell.value)
    return result


def _description_merged_ranges(
    sheet: Worksheet, num_description_rows: int
) -> list[str]:
    """Merged range refs (e.g. 'A1:D1') in the first num_description_rows."""
    result: list[str] = []
    for merged_range in sheet.merged_cells.ranges:
        if merged_range.max_row <= num_description_rows:
            result.append(str(merged_range))
    return result


def _read_sheet_with_structure(
    excel_file: pd.ExcelFile,
    sheet_name: str,
    header_row: int = 0,
    input_file: str | None = None,
    skip_rows_after_header: list[int] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, list[list[Any]], list[str]]:
    """
    Read a sheet preserving description rows and structure.

    Uses openpyxl to resolve merged cells in the header row so column names
    are correct when the sheet has merged cells. Rows whose indices are in
    skip_rows_after_header (0-based) are excluded from the data (e.g. a
    data-type row immediately below the header).

    Returns:
        Tuple of (data_df, description_df, description_rows,
        description_merged_ranges)
        - data_df: DataFrame with header row as column names and data rows
        - description_df: DataFrame with description rows (if any)
        - description_rows: List of description row data (for writing back)
        - description_merged_ranges: Merged range refs to preserve when writing
    """
    # Read entire sheet without header to preserve all rows
    full_df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)
    max_col = full_df.shape[1]
    description_merged_ranges: list[str] = []

    # Resolve header row (and optional description merged ranges) via openpyxl
    if input_file and max_col:
        wb = load_workbook(input_file, read_only=False, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Header row in 1-based openpyxl
            header_row_1based = header_row + 1
            header_values = _get_row_values_resolving_merged(
                ws, header_row_1based, max_col
            )
            # Pandas-friendly column names (no None)
            columns = []
            for i, v in enumerate(header_values):
                if v is None or (
                    isinstance(v, float) and pd.isna(cast("float", v))
                ):
                    columns.append(f"Unnamed: {i}")
                else:
                    columns.append(str(v).strip() or f"Unnamed: {i}")

            if header_row > 0:
                description_merged_ranges = _description_merged_ranges(
                    ws, header_row
                )

            # Description rows (rows before header)
            description_rows = (
                full_df.iloc[:header_row].values.tolist()
                if header_row > 0
                else []
            )
            # Data rows: everything after header, optionally excluding some rows
            data_block = full_df.iloc[header_row + 1 :]
            if skip_rows_after_header:
                # Drop by 0-based index (data_block index = original row)
                to_drop = [
                    i
                    for i in data_block.index
                    if i in skip_rows_after_header
                ]
                data_block = data_block.drop(index=to_drop, errors="ignore")

            data_df = pd.DataFrame(data_block.values, columns=columns)
            description_df = (
                full_df.iloc[:header_row].copy()
                if header_row > 0
                else pd.DataFrame()
            )
            wb.close()
            return (
                data_df,
                description_df,
                description_rows,
                description_merged_ranges,
            )

    # Fallback when no openpyxl path or empty sheet
    if header_row == 0:
        description_rows = []
        description_df = pd.DataFrame()
        data_df = pd.read_excel(
            excel_file, sheet_name=sheet_name, header=0
        )
        if skip_rows_after_header:
            data_df = data_df.drop(index=skip_rows_after_header, errors="ignore")
    else:
        description_rows = full_df.iloc[:header_row].values.tolist()
        description_df = full_df.iloc[:header_row].copy()
        data_df = pd.read_excel(
            excel_file, sheet_name=sheet_name, header=header_row
        )
        if skip_rows_after_header:
            data_df = data_df.drop(index=skip_rows_after_header, errors="ignore")

    return (
        data_df,
        description_df,
        description_rows,
        description_merged_ranges,
    )


def _write_sheet_with_structure(
    writer: pd.ExcelWriter,
    sheet_name: str,
    data_df: pd.DataFrame,
    description_rows: list[list[Any]],
    header_row: int,
    date_columns: list[str] | None = None,
    date_format: str | None = None,
    description_merged_ranges: list[str] | None = None,
) -> None:
    """
    Write a sheet preserving description rows and structure.

    Args:
        writer: ExcelWriter instance.
        sheet_name: Name of the sheet to write.
        data_df: DataFrame with data to write.
        description_rows: List of description rows to write at the top.
        header_row: Row index where header should be written.
        date_columns: Optional list of date column names to format.
        date_format: Optional Excel date format string (e.g., 'yyyy-mm-dd').
        description_merged_ranges: Optional merged cell refs to restore
            (e.g. 'A1:D1').
    """
    # Convert Python date format to Excel format
    excel_date_format = None
    if date_format:
        # Convert common Python format strings to Excel format
        # Excel uses lowercase: yyyy (year), mm (month), dd (day)
        excel_date_format = (
            date_format.replace("YYYY", "yyyy")
            .replace("YY", "yy")
            .replace("DD", "dd")
            .replace("MM", "mm")
        )

    # Calculate where to start writing data (after description rows + header row)
    data_start_row = len(description_rows) + 1 if description_rows else 1

    # Write data without header first (header=False), then we'll add header manually
    data_df.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        header=False,
        startrow=data_start_row,
    )

    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = workbook[sheet_name]

    # Write description rows at the top
    if description_rows:
        for row_idx, desc_row in enumerate(description_rows, start=1):
            for col_idx, value in enumerate(desc_row, start=1):
                cell_value = value
                # Handle NaN values
                if pd.isna(cell_value):
                    cell_value = None
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
        # Restore merged cells in the description area
        if description_merged_ranges:
            for range_ref in description_merged_ranges:
                with contextlib.suppress(Exception):
                    worksheet.merge_cells(range_ref)

    # Write header row (after description rows)
    header_row_idx = len(description_rows) + 1 if description_rows else 1
    for col_idx, col_name in enumerate(data_df.columns, start=1):
        worksheet.cell(row=header_row_idx, column=col_idx, value=col_name)

    # Apply date formatting to date columns if specified
    if excel_date_format and date_columns:
        # Find column indices for date columns
        col_map = {col: idx + 1 for idx, col in enumerate(data_df.columns)}
        for date_col in date_columns:
            if date_col in col_map:
                col_idx = col_map[date_col]
                # Apply format to all data rows (skip description and header rows)
                for row_idx in range(
                    data_start_row + 1, data_start_row + len(data_df) + 1
                ):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    # Only format if cell has a date value
                    if cell.value is not None:
                        cell.number_format = excel_date_format
