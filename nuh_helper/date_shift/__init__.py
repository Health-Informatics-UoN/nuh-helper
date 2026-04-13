"""
Date shifting for patient data in Excel spreadsheets.

Consistently shifts dates for patient IDs across multiple sheets and columns
in an Excel file, with support for reproducible shifts using a linking table.
"""

import logging
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any, cast

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from nuh_helper.date_shift import _excel, _parse, mappings

logger = logging.getLogger(__name__)


def _get_patient_ids_and_shift_mappings(
    input_file: str,
    patient_sheet: str,
    patient_id_col: str,
    sheet_configs: dict[str, dict[str, Any]],
    min_shift_days: int,
    max_shift_days: int,
    linking_table_path: str | None,
    seed: int | None,
    patient_header_row: int,
    patient_skip_rows: list[int] | None,
) -> tuple[list[str], pd.DataFrame]:
    """
    Read patient IDs from the patient sheet and resolve shift mappings
    (load from CSV or generate). Returns (patient_ids, shift_mappings)
    with shift_mappings already normalized and deduplicated.
    """
    effective_patient_header_row = patient_header_row
    effective_patient_skip_rows = patient_skip_rows
    if patient_sheet in sheet_configs:
        cfg = sheet_configs[patient_sheet]
        effective_patient_header_row = cast(
            int, cfg.get("header_row", patient_header_row)
        )
        effective_patient_skip_rows = cfg.get(
            "skip_rows_after_header", patient_skip_rows
        )

    patient_excel = pd.ExcelFile(input_file, engine="openpyxl")
    patient_df, _, _, _ = _excel._read_sheet_with_structure(
        patient_excel,
        sheet_name=patient_sheet,
        header_row=effective_patient_header_row,
        input_file=input_file,
        skip_rows_after_header=effective_patient_skip_rows,
    )
    if patient_id_col not in patient_df.columns:
        raise ValueError(
            f"Patient ID column '{patient_id_col}' not found in sheet '{patient_sheet}'"
        )

    patient_ids = (
        patient_df[patient_id_col]
        .apply(_parse._normalize_patient_id)
        .dropna()
        .unique()
        .tolist()
    )
    logger.info("Found %d patient(s) in sheet '%s'", len(patient_ids), patient_sheet)

    if linking_table_path and Path(linking_table_path).exists():
        logger.info("Loading shift mappings from '%s'", linking_table_path)
        shift_mappings = mappings.load_shift_mappings(linking_table_path)
        shift_mappings = shift_mappings[shift_mappings["patient_id"].isin(patient_ids)]
        existing_ids = set(shift_mappings["patient_id"])
        missing_ids = [pid for pid in patient_ids if pid not in existing_ids]
        if missing_ids:
            logger.warning(
                "%d patient(s) had no entry in the linking table; new shifts generated",
                len(missing_ids),
            )
            new_shifts = mappings.generate_shift_mappings(
                missing_ids, min_shift_days, max_shift_days, seed
            )
            shift_mappings = pd.concat([shift_mappings, new_shifts], ignore_index=True)
    else:
        logger.info("Generating shift mappings for %d patient(s)", len(patient_ids))
        shift_mappings = mappings.generate_shift_mappings(
            patient_ids, min_shift_days, max_shift_days, seed
        )

    shift_mappings["patient_id"] = shift_mappings["patient_id"].apply(
        _parse._normalize_patient_id
    )
    shift_mappings = shift_mappings.dropna(subset=["patient_id"]).drop_duplicates(
        subset=["patient_id"], keep="first"
    )
    return (patient_ids, shift_mappings)


def apply_date_shifts(
    df: pd.DataFrame,
    patient_id_col: str,
    date_columns: list[str],
    shift_mappings: pd.DataFrame,
    date_format: str | None = None,
    shift_exceptions: dict[str, list[str]] | None = None,
) -> pd.DataFrame:
    """
    Apply date shifts to specified columns in a DataFrame.

    Args:
        df: pd.DataFrame containing patient data.
        patient_id_col: Name of the column containing patient IDs.
        date_columns: List of column names containing dates to shift.
        shift_mappings: DataFrame with 'patient_id' and 'shift_days' columns.
        date_format: Optional date format string (e.g., 'YYYY-MM-DD').
                     Note: This parameter is kept for API compatibility but formatting
                     is applied at the Excel cell level, not in the DataFrame.
        shift_exceptions: Optional dict mapping column names to lists of date strings
                          that should never be shifted (e.g. a fixed end-of-study date).

    Returns:
        DataFrame with shifted dates.
    """
    df = df.copy()

    # Normalize patient IDs in the working DataFrame to align with mapping keys
    df[patient_id_col] = df[patient_id_col].apply(_parse._normalize_patient_id)

    shift_dict = dict(
        zip(
            shift_mappings["patient_id"],
            shift_mappings["shift_days"],
            strict=True,
        )
    )

    # Pre-parse exception dates once per column
    parsed_exceptions: dict[str, set[date]] = {}
    if shift_exceptions:
        for col, exc_values in shift_exceptions.items():
            parsed_set: set[date] = set()
            for v in exc_values:
                ts = _parse._parse_date_value(v)
                if ts is not None:
                    parsed_set.add(ts.date())
            if parsed_set:
                parsed_exceptions[col] = parsed_set

    for date_col in date_columns:
        if date_col not in df.columns:
            logger.warning(
                "Date column '%s' not found in DataFrame, skipping", date_col
            )
            continue

        # Parse flexible date strings (handles YYYY-DD-MM and placeholders "Unknown")
        non_null_before = df[date_col].notna().sum()
        df[date_col] = df[date_col].apply(_parse._parse_date_value)
        parse_failures = non_null_before - sum(x is not None for x in df[date_col])
        if parse_failures > 0:
            logger.debug(
                "Column '%s': %d value(s) could not be parsed as dates",
                date_col,
                parse_failures,
            )

        # Apply shifts
        exc_dates = parsed_exceptions.get(date_col, set())
        df[date_col] = df.apply(
            lambda row: (
                row[date_col]  # noqa: B023
                + pd.Timedelta(days=shift_dict.get(row[patient_id_col], 0))
                if row[date_col] is not None  # noqa: B023
                and row[patient_id_col] in shift_dict  # noqa: B023
                and row[date_col].date() not in exc_dates  # noqa: B023
                else row[date_col]  # noqa: B023
            ),
            axis=1,
        )

        # Convert back to date-only format (removes time component)
        df[date_col] = df[date_col].apply(
            lambda x: (
                x.date() if isinstance(x, (pd.Timestamp, datetime, date)) else None
            ),
        )

    return df


def shift_excel_dates(
    input_file: str,
    output_file: str,
    patient_sheet: str,
    patient_id_col: str,
    sheet_configs: dict[str, dict[str, Any]],
    min_shift_days: int = -15,
    max_shift_days: int = 15,
    linking_table_path: str | None = None,
    linking_table_output: str | None = None,
    seed: int | None = None,
    patient_header_row: int = 0,
    patient_skip_rows: list[int] | None = None,
    date_format: str | None = None,
) -> None:
    """
    Shift dates in an Excel file for patient IDs consistently across sheets.

    Args:
        input_file: Path to input Excel file.
        output_file: Path to output Excel file with shifted dates.
        patient_sheet: Name of the sheet containing patient IDs.
        patient_id_col: Name of the column containing patient IDs in the patient sheet.
        sheet_configs: Dictionary mapping sheet names to configuration dicts.
                      Each config dict should have:
                      - 'patient_id_col': Name of patient ID column in that sheet
                      - 'date_columns': List of date column names to shift
                      Optional per-sheet header handling:
                      - 'header_row': zero-based row index of the column names (default 0)
                      - 'skip_rows_after_header': list of zero-based row indices to exclude
                        from data (e.g. a data-type row immediately below the header)
        min_shift_days: Minimum number of days to shift (default: -15).
        max_shift_days: Maximum number of days to shift (default: 15).
        linking_table_path: Optional path to existing linking table CSV for reproducibility.
        linking_table_output: Path to save the linking table CSV (default: 'shift_mappings.csv').
        seed: Optional random seed for generating shifts.
        patient_header_row: Zero-based header row index for the patient sheet (default: 0).
        patient_skip_rows: Optional zero-based row indices to exclude from patient data
                     (e.g. a data-type row immediately below the header).
        date_format: Optional Excel date format string (e.g., 'YYYY-MM-DD', 'yyyy-mm-dd').
                     If None, Excel's default date format is used.
                     Common formats: 'YYYY-MM-DD', 'MM/DD/YYYY', 'DD-MM-YYYY', etc.
    """  # noqa: E501
    logger.info("Shifting dates: '%s' → '%s'", input_file, output_file)
    logger.debug(
        "Shift range: %d to %d days, seed=%s",
        min_shift_days,
        max_shift_days,
        seed,
    )

    _patient_ids, shift_mappings = _get_patient_ids_and_shift_mappings(
        input_file=input_file,
        patient_sheet=patient_sheet,
        patient_id_col=patient_id_col,
        sheet_configs=sheet_configs,
        min_shift_days=min_shift_days,
        max_shift_days=max_shift_days,
        linking_table_path=linking_table_path,
        seed=seed,
        patient_header_row=patient_header_row,
        patient_skip_rows=patient_skip_rows,
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        excel_file = pd.ExcelFile(input_file, engine="openpyxl")

        for sheet_name in excel_file.sheet_names:
            default_header_row = 0
            header_row = default_header_row
            sheet_date_columns: list[str] | None = None
            skip_rows_after_header: list[int] | None = None
            sheet_shift_exceptions: dict[str, list[str]] | None = None

            if sheet_name in sheet_configs:
                config = sheet_configs[cast(str, sheet_name)]
                sheet_patient_id_col: str = cast(str, config["patient_id_col"])
                date_columns: list[str] = cast(list[str], config["date_columns"])
                header_row = cast(int, config.get("header_row", header_row))
                skip_rows_after_header = config.get("skip_rows_after_header")
                sheet_shift_exceptions = config.get("shift_exceptions")
                sheet_date_columns = date_columns
                logger.info(
                    "Shifting %d date column(s) in sheet '%s'",
                    len(date_columns),
                    sheet_name,
                )

            df, _description_df, description_rows, description_merged_ranges = (
                _excel._read_sheet_with_structure(
                    excel_file,
                    sheet_name=cast(str, sheet_name),
                    header_row=header_row,
                    input_file=input_file,
                    skip_rows_after_header=skip_rows_after_header,
                )
            )

            if sheet_name in sheet_configs:
                if sheet_patient_id_col not in df.columns:
                    raise ValueError(
                        f"Patient ID column '{sheet_patient_id_col}' not found in sheet '{sheet_name}'"  # noqa: E501
                    )

                df = apply_date_shifts(
                    df,
                    sheet_patient_id_col,
                    date_columns,
                    shift_mappings,
                    date_format=None,
                    shift_exceptions=sheet_shift_exceptions,
                )

            _excel._write_sheet_with_structure(
                writer,
                sheet_name=cast(str, sheet_name),
                data_df=df,
                description_rows=description_rows,
                header_row=header_row,
                date_columns=sheet_date_columns,
                date_format=date_format,
                description_merged_ranges=description_merged_ranges,
            )

    logger.info("Output written to '%s'", output_file)

    linking_path = linking_table_output or "shift_mappings.csv"
    shift_mappings.to_csv(linking_path, index=False)
    logger.info("Linking table saved to '%s'", linking_path)


def shift_excel_dates_inplace(
    input_file: str,
    output_file: str,
    patient_sheet: str,
    patient_id_col: str,
    sheet_configs: dict[str, dict[str, Any]],
    min_shift_days: int = -15,
    max_shift_days: int = 15,
    linking_table_path: str | None = None,
    linking_table_output: str | None = None,
    seed: int | None = None,
    patient_header_row: int = 0,
    patient_skip_rows: list[int] | None = None,
) -> None:
    """
    Shift dates in an Excel file, preserving all cell formatting.

    Unlike shift_excel_dates(), this function copies the input file and then
    modifies date cells directly via openpyxl, so all formatting (cell styles,
    merged cells, column widths, conditional formatting, etc.) is preserved.

    Args:
        input_file: Path to input Excel file.
        output_file: Path for the output file (copy of input with shifted dates).
        patient_sheet: Name of the sheet containing patient IDs.
        patient_id_col: Name of the column containing patient IDs in the patient sheet.
        sheet_configs: Dictionary mapping sheet names to configuration dicts.
                      Each config dict should have:
                      - 'patient_id_col': Name of patient ID column in that sheet
                      - 'date_columns': List of date column names to shift
                      Optional per-sheet header handling:
                      - 'header_row': zero-based row index of the column names (default 0)
                      - 'skip_rows_after_header': list of zero-based row indices to exclude
                        from data (e.g. a data-type row immediately below the header)
        min_shift_days: Minimum number of days to shift (default: -15).
        max_shift_days: Maximum number of days to shift (default: 15).
        linking_table_path: Optional path to existing linking table CSV for reproducibility.
        linking_table_output: Path to save the linking table CSV (default: 'shift_mappings.csv').
        seed: Optional random seed for generating shifts.
        patient_header_row: Zero-based header row index for the patient sheet (default: 0).
        patient_skip_rows: Optional zero-based row indices to exclude from patient data.
    """  # noqa: E501
    logger.info("Shifting dates in-place: '%s' → '%s'", input_file, output_file)
    logger.debug(
        "Shift range: %d to %d days, seed=%s",
        min_shift_days,
        max_shift_days,
        seed,
    )

    shutil.copy2(input_file, output_file)

    _patient_ids, shift_mappings = _get_patient_ids_and_shift_mappings(
        input_file=input_file,
        patient_sheet=patient_sheet,
        patient_id_col=patient_id_col,
        sheet_configs=sheet_configs,
        min_shift_days=min_shift_days,
        max_shift_days=max_shift_days,
        linking_table_path=linking_table_path,
        seed=seed,
        patient_header_row=patient_header_row,
        patient_skip_rows=patient_skip_rows,
    )

    shift_dict: dict[str, int] = dict(
        zip(
            shift_mappings["patient_id"],
            shift_mappings["shift_days"],
            strict=True,
        )
    )

    wb = load_workbook(output_file, keep_links=False)
    wb.defined_names.clear()

    for sheet_name, config in sheet_configs.items():
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook, skipping", sheet_name)
            continue

        ws = cast(Worksheet, wb[sheet_name])
        sheet_patient_id_col: str = cast(str, config["patient_id_col"])
        date_columns: list[str] = cast(list[str], config["date_columns"])
        header_row: int = cast(int, config.get("header_row", 0))
        skip_rows_after_header: list[int] | None = config.get("skip_rows_after_header")

        max_col = ws.max_column or 0
        if not max_col:
            continue

        header_row_1based = header_row + 1
        header_values = _excel._get_row_values_resolving_merged(
            ws, header_row_1based, max_col
        )
        col_index: dict[str, int] = {}
        for i, val in enumerate(header_values, start=1):
            if val is not None and str(val).strip():
                col_index[str(val).strip()] = i

        if sheet_patient_id_col not in col_index:
            raise ValueError(
                f"Patient ID column '{sheet_patient_id_col}' not found in sheet '{sheet_name}'"  # noqa: E501
            )

        pid_col_idx = col_index[sheet_patient_id_col]
        date_col_indices: dict[str, int] = {}
        for col in date_columns:
            if col in col_index:
                date_col_indices[col] = col_index[col]
            else:
                logger.warning(
                    "Date column '%s' not found in sheet '%s', skipping",
                    col,
                    sheet_name,
                )

        if not date_col_indices:
            continue

        # Pre-parse exception dates once per column
        parsed_exceptions: dict[str, set[date]] = {}
        shift_exceptions_config: dict[str, list[str]] | None = config.get(
            "shift_exceptions"
        )
        if shift_exceptions_config:
            for exc_col, exc_values in shift_exceptions_config.items():
                parsed_set: set[date] = set()
                for v in exc_values:
                    ts = _parse._parse_date_value(v)
                    if ts is not None:
                        parsed_set.add(ts.date())
                if parsed_set:
                    parsed_exceptions[exc_col] = parsed_set

        skip_row_set: set[int] = (
            {idx + 1 for idx in skip_rows_after_header}
            if skip_rows_after_header
            else set()
        )

        logger.info(
            "Shifting %d date column(s) in sheet '%s'",
            len(date_col_indices),
            sheet_name,
        )

        data_start_1based = header_row + 2
        for row_idx in range(data_start_1based, (ws.max_row or 0) + 1):
            if row_idx in skip_row_set:
                continue

            pid_cell = ws.cell(row=row_idx, column=pid_col_idx)
            pid = _parse._normalize_patient_id(pid_cell.value)
            shift_days = shift_dict.get(pid) if pid is not None else None

            for col_name, date_col_idx in date_col_indices.items():
                cell = ws.cell(row=row_idx, column=date_col_idx)
                original_value = cell.value
                parsed = _parse._parse_date_value(original_value)

                if parsed is None:
                    if original_value is not None and not isinstance(
                        original_value, (datetime, date)
                    ):
                        cell.value = None
                    continue

                if shift_days is None:
                    continue

                exc_dates = parsed_exceptions.get(col_name, set())
                if exc_dates and parsed.date() in exc_dates:
                    continue

                shifted = parsed + pd.Timedelta(days=shift_days)
                if isinstance(original_value, date) and not isinstance(
                    original_value, datetime
                ):
                    cell.value = cast(Any, shifted.to_pydatetime().date())
                else:
                    cell.value = cast(Any, shifted.to_pydatetime())

    wb.save(output_file)
    logger.info("Output written to '%s'", output_file)

    linking_path = linking_table_output or "shift_mappings.csv"
    shift_mappings.to_csv(linking_path, index=False)
    logger.info("Linking table saved to '%s'", linking_path)


# Re-export public API
from nuh_helper.date_shift.mappings import (  # noqa: E402
    generate_shift_mappings,
    load_shift_mappings,
)

__all__ = [
    "shift_excel_dates",
    "shift_excel_dates_inplace",
    "apply_date_shifts",
    "generate_shift_mappings",
    "load_shift_mappings",
]
