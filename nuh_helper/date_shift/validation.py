"""This contains an inspection function and error record type to determine if a
spreadsheet has data in abnormal places. it's mean tot check for "little notes" which
are outside of the CDM and may have undocumented patient data"""

from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path


class Error:
    """base class for the errors. has a simplified __eq__ for `assert error in list`"""

    def __eq__(self, them: object) -> bool:
        if type(self) is not type(them):
            return False
        return str(self) == str(them)


class ExcessRows(Error):
    """error indicating that there are extra rows in a spreadsheet that don't have a
    patient id and won't be shifted"""

    def __init__(self, sheet_name: str, excess: list[int]) -> None:
        self.sheet_name = sheet_name
        self.excess = excess

    def __str__(self) -> str:
        return f"ExcessRows('{self.sheet_name}', {self.excess})"


class UnlabeledColumns(Error):
    """indication that there are columns with data but no header; probably notes in the
    margin about missing tests or (previously) dates related to patient's treatment to
    explain the data in the spreadsheet."""

    def __init__(self, sheet_name: str, columns: list[int]) -> None:
        self.sheet_name = sheet_name
        self.columns = columns

    def __str__(self) -> str:
        return f"UnlabeledColumns('{self.sheet_name}', {self.columns})"


class PatientColumnMissing(Error):
    """used to indicate that the patien column wasn't found in the spreadsheet"""

    def __init__(self, sheet_name: str, label: str) -> None:
        self.sheet_name = sheet_name
        self.name = label

    def __str__(self) -> str:
        return f"PatientColumnMissing('{self.sheet_name}', '{self.label}')"


def format_errors(errors: list[Error]) -> str:
    """formats a collection of error objects into a human digestible string"""
    message: str = ""
    names = []

    # group the errors by sheet names
    for error in errors:
        if error.sheet_name not in names:
            names.append(error.sheet_name)

    for sheet_name in names:
        message += f"on sheet {sheet_name=} ...\n"
        for error in errors:
            if error.sheet_name != sheet_name:
                continue
            match error:
                case ExcessRows():
                    message += (
                        f"\tthere were {len(error.excess)} rows with data but no "
                        + "patient ID\n"
                    )
                    message += f"\t\t{error.excess}\n"
                case UnlabeledColumns():
                    message += (
                        f"\tthere were {len(error.columns)} columns with no data "
                        + "in their label\n"
                    )
                    message += f"\t\t{error.columns}\n"
                case PatientColumnMissing():
                    label = error.label
                    message += f"\tthere was no patient column {label=}\n"
    return message


def inspect(sheet_file: Path, sheet_configs: dict) -> list[Error]:
    """Find data that's out of bounds in the spreadsheet. Uses the date-shifting
    sheet_configs structure. Rather than throw exceptions, this returns a list of Error
    objects that can be inspected or tested for."""

    from openpyxl import load_workbook

    errors: list[Error] = []

    workbook = load_workbook(sheet_file, read_only=True, rich_text=False)
    for sheet_name in workbook.sheetnames:
        if sheet_name not in sheet_configs:
            print(f"skipping sheet {sheet_name=} since there's no config for it")
            continue

        sheet = workbook[sheet_name]

        # scan the header row to find out what the bounds of the spreadsheet should be
        header_row = sheet_configs[sheet_name]["header_row"]
        patient_id_col_text = sheet_configs[sheet_name]["patient_id_col"]
        skip_rows = sheet_configs[sheet_name]["skip_rows_after_header"]

        # we'll want to use the index in later checks
        patient_id_col_index: None | int = None

        # record the "blank" columns in the 
        blanks: list[int] = []

        # check each cell of the header
        for col in range(0, sheet.max_column):
            value = sheet.cell(header_row + 1, col + 1)
            if blank_cell(value):
                blanks.append(col)
            elif value.value == patient_id_col_text:
                patient_id_col_index = col

        if blanks:
            errors.append(UnlabeledColumns(sheet_name, blanks))

        # we can't do any further checks without the patient_id_col_index
        if patient_id_col_index is None:
            errors.append(PatientColumnMissing(sheet_name, patient_id_col_text))
        else:
            excess = []

            # find any rows with data but no patient id
            for row in range(0, sheet.max_row):
                if row in skip_rows or row == header_row:
                    continue

                # we will allow "blank" rows
                # ... such as empty rows between groups of patients
                should_be_blank = blank_cell(
                    sheet.cell(row + 1, patient_id_col_index + 1)
                )

                # to allow "whitespace rows" we only check rows without a patient id
                if should_be_blank and not blank_row(sheet, row):
                    excess.append(row)

            if excess:
                errors.append(ExcessRows(sheet_name, excess))

    return errors


def blank_cell(cell: Cell) -> bool:
    """tests if a cell value is blank"""
    return str(cell.value).strip() == "" or cell.value is None


def blank_row(sheet: Worksheet, row: int) -> bool:
    """tests if a row of a Worksheet is blank"""
    for c in range(0, sheet.max_column):
        cell = sheet.cell(row + 1, c + 1)
        if not blank_cell(cell):
            return False
    return True
