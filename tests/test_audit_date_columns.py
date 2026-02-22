"""Tests for audit_date_columns."""

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from nuh_helper import audit_date_columns

DATA_DIR = Path(__file__).parent / "data"
INPUT_FILE = DATA_DIR / "patients.xlsx"

# Full correct config for patients.xlsx
FULL_CONFIG = {
    "patients": {
        "patient_id_col": "patient_id",
        "date_columns": ["dob", "last_alive"],
    },
    "results": {
        "patient_id_col": "patient_id",
        "date_columns": ["date_result"],
    },
}


class TestNoFindings:
    def test_empty_dict_when_all_date_columns_configured(self) -> None:
        result = audit_date_columns(str(INPUT_FILE), FULL_CONFIG)
        assert result == {}

    def test_empty_dict_when_no_date_columns_exist(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws["A1"] = "patient_id"
        ws["B1"] = "name"
        ws["C1"] = "age"
        ws["A2"] = "P001"
        ws["B2"] = "Alice"
        ws["C2"] = 30
        path = tmp_path / "no_dates.xlsx"
        wb.save(path)

        result = audit_date_columns(
            str(path),
            {"data": {"patient_id_col": "patient_id", "date_columns": []}},
        )
        assert result == {}


class TestFindings:
    def test_detects_unconfigured_date_column_on_configured_sheet(self) -> None:
        # Omit "dob" from the config — it should be flagged
        partial_config = {
            "patients": {
                "patient_id_col": "patient_id",
                "date_columns": ["last_alive"],  # dob intentionally missing
            },
            "results": {
                "patient_id_col": "patient_id",
                "date_columns": ["date_result"],
            },
        }
        result = audit_date_columns(str(INPUT_FILE), partial_config)
        assert "patients" in result
        assert "dob" in result["patients"]

    def test_detects_date_columns_on_unconfigured_sheet(
        self, tmp_path: Path
    ) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "events"
        ws["A1"] = "patient_id"
        ws["B1"] = "event_date"
        ws["A2"] = "P001"
        ws["B2"] = date(2023, 1, 15)
        ws["A3"] = "P002"
        ws["B3"] = date(2023, 6, 1)
        path = tmp_path / "events.xlsx"
        wb.save(path)

        # Pass an empty sheet_configs — no sheets configured at all
        result = audit_date_columns(str(path), {})
        assert "events" in result
        assert "event_date" in result["events"]

    def test_does_not_flag_patient_id_column(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws["A1"] = "patient_id"
        ws["B1"] = "visit_date"
        ws["A2"] = "P001"
        ws["B2"] = date(2023, 1, 15)
        path = tmp_path / "simple.xlsx"
        wb.save(path)

        result = audit_date_columns(
            str(path),
            {"data": {"patient_id_col": "patient_id", "date_columns": ["visit_date"]}},
        )
        assert result == {}

    def test_does_not_flag_configured_date_columns(self, tmp_path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws["A1"] = "patient_id"
        ws["B1"] = "visit_date"
        ws["A2"] = "P001"
        ws["B2"] = date(2023, 1, 15)
        path = tmp_path / "simple.xlsx"
        wb.save(path)

        result = audit_date_columns(
            str(path),
            {"data": {"patient_id_col": "patient_id", "date_columns": ["visit_date"]}},
        )
        assert result == {}


class TestThreshold:
    def _make_workbook_with_mixed_column(
        self, tmp_path: Path, date_count: int, total: int
    ) -> Path:
        """Sheet with `date_count` dates and `total - date_count` strings."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws["A1"] = "patient_id"
        ws["B1"] = "maybe_dates"
        for i in range(total):
            ws.cell(row=i + 2, column=1, value=f"P{i:03d}")
            if i < date_count:
                ws.cell(row=i + 2, column=2, value=date(2023, 1, i % 28 + 1))
            else:
                ws.cell(row=i + 2, column=2, value="not a date")
        path = tmp_path / "mixed.xlsx"
        wb.save(path)
        return path

    def test_column_above_threshold_is_flagged(self, tmp_path: Path) -> None:
        # 8 out of 10 = 0.8 ratio → flagged at default threshold of 0.5
        path = self._make_workbook_with_mixed_column(tmp_path, date_count=8, total=10)
        result = audit_date_columns(
            str(path),
            {"data": {"patient_id_col": "patient_id", "date_columns": []}},
        )
        assert "data" in result
        assert "maybe_dates" in result["data"]

    def test_column_below_threshold_is_not_flagged(self, tmp_path: Path) -> None:
        # 3 out of 10 = 0.3 ratio → not flagged at default threshold of 0.5
        path = self._make_workbook_with_mixed_column(tmp_path, date_count=3, total=10)
        result = audit_date_columns(
            str(path),
            {"data": {"patient_id_col": "patient_id", "date_columns": []}},
        )
        assert result == {}

    def test_custom_threshold_lowers_sensitivity(self, tmp_path: Path) -> None:
        # 3/10 = 0.3 → flagged when threshold=0.2
        path = self._make_workbook_with_mixed_column(tmp_path, date_count=3, total=10)
        result = audit_date_columns(
            str(path),
            {"data": {"patient_id_col": "patient_id", "date_columns": []}},
            threshold=0.2,
        )
        assert "data" in result
        assert "maybe_dates" in result["data"]

    def test_zero_date_values_never_flagged(self, tmp_path: Path) -> None:
        path = self._make_workbook_with_mixed_column(tmp_path, date_count=0, total=5)
        result = audit_date_columns(
            str(path),
            {"data": {"patient_id_col": "patient_id", "date_columns": []}},
            threshold=0.0,
        )
        assert result == {}
