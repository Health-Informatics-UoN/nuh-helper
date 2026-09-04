import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from nuh_helper import shift_excel_dates_inplace


@pytest.mark.parametrize("clamp", [True, False])
def test_clamp_dates(clamp: bool, tmp_path: Path) -> None:

    clamp_date: None | datetime.datetime = None
    if clamp:
        clamp_date = datetime.datetime(2024, 9, 18)

    source_file = Path(__file__).parent / "data/clamped/Book5.xlsx"
    output_path = tmp_path / "target.xlsx"
    linking_table_old = Path(__file__).parent / "data/clamped/book-5.csv"
    linking_table_out = tmp_path / "linking_table_out.csv"

    sheet_configs = {
        "people": {
            "patient_id_col": "ptid",
            "date_columns": ["dob"],
            "header_row": 1,
            "skip_rows_after_header": [],
        },
        "events": {
            "patient_id_col": "patient",
            "date_columns": ["date"],
            "header_row": 0,
            "skip_rows_after_header": [],
        },
    }

    shift_excel_dates_inplace(
        input_file=str(source_file),
        output_file=str(output_path),
        patient_sheet="people",
        patient_id_col="ptid",
        sheet_configs=sheet_configs,
        min_shift_days=-20,
        max_shift_days=-1,
        seed=14333,
        linking_table_path=str(linking_table_old),
        linking_table_output=str(linking_table_out),
        clamp_date=clamp_date,
    )

    workbook = load_workbook(output_path)

    # check the first sheet
    people = workbook.worksheets[0]

    assert people.cell(1, 1).value == "person id"
    assert people.cell(2, 1).value == "ptid"
    assert people.cell(3, 1).value == "nuh17"
    assert people.cell(4, 1).value == "nuh28"
    assert people.cell(5, 1).value == "nuh71"

    assert people.cell(1, 2).value == "date of birth"
    assert people.cell(2, 2).value == "dob"
    assert str(people.cell(3, 2).value) == "1990-08-18 00:00:00"
    assert str(people.cell(4, 2).value) == "1987-09-10 00:00:00"
    assert str(people.cell(5, 2).value) == "1976-02-19 00:00:00"

    assert people.cell(1, 3).value == "something else"
    assert people.cell(2, 3).value is None
    assert people.cell(3, 3).value == "tacos"
    assert people.cell(4, 3).value == "pizza"
    assert people.cell(5, 3).value == "cake"

    # check the first sheet
    people = workbook.worksheets[0]

    assert people.cell(1, 1).value == "person id"
    assert people.cell(2, 1).value == "ptid"
    assert people.cell(3, 1).value == "nuh17"
    assert people.cell(4, 1).value == "nuh28"
    assert people.cell(5, 1).value == "nuh71"

    assert people.cell(1, 2).value == "date of birth"
    assert people.cell(2, 2).value == "dob"
    assert str(people.cell(3, 2).value) == "1990-08-18 00:00:00"
    assert str(people.cell(4, 2).value) == "1987-09-10 00:00:00"
    assert str(people.cell(5, 2).value) == "1976-02-19 00:00:00"

    assert people.cell(1, 3).value == "something else"
    assert people.cell(2, 3).value is None
    assert people.cell(3, 3).value == "tacos"
    assert people.cell(4, 3).value == "pizza"
    assert people.cell(5, 3).value == "cake"

    # check the events sheet
    events = workbook.worksheets[1]

    assert events.cell(1, 1).value == "patient"
    assert events.cell(2, 1).value == "nuh17"
    assert events.cell(3, 1).value is None
    assert events.cell(4, 1).value == "nuh28"
    assert events.cell(5, 1).value == "nuh28"
    assert events.cell(6, 1).value is None
    assert events.cell(7, 1).value == "nuh71"

    # columns 2 and 4 are fully blank
    for i in range(7):
        assert events.cell(i + 1, 2).value is None
        assert events.cell(i + 1, 4).value is None

    # column 5 is not dates so "easy"
    assert events.cell(1, 5).value == "pill"
    assert events.cell(2, 5).value == "asprin"
    assert events.cell(3, 5).value is None
    assert str(events.cell(4, 5).value) == "2027-06-19 00:00:00"
    assert events.cell(5, 5).value == "zuul"
    assert events.cell(6, 5).value is None
    assert events.cell(7, 5).value == "ibuprophen"

    # column 3 will have the clamped dates
    assert events.cell(1, 3).value == "date"
    assert str(events.cell(2, 3).value) == (
        "2025-06-06 00:00:00" if not clamp else str(clamp_date)
    )
    assert events.cell(3, 3).value is None
    assert str(events.cell(4, 3).value) == (
        "2025-08-31 00:00:00" if not clamp else str(clamp_date)
    )
    assert str(events.cell(5, 3).value) == (
        "2026-07-12 00:00:00" if not clamp else str(clamp_date)
    )
    assert events.cell(6, 3).value is None
    assert str(events.cell(7, 3).value) == (
        "2025-09-16 00:00:00" if not clamp else str(clamp_date)
    )
