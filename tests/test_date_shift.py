"""Unit tests for nuh_helper.date_shift"""
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from nuh_helper.date_shift import (
    apply_date_shifts,
    generate_shift_mappings,
    load_shift_mappings,
    shift_excel_dates,
)
from nuh_helper.date_shift._excel import (
    _description_merged_ranges,
    _get_row_values_resolving_merged,
)
from nuh_helper.date_shift._parse import _normalize_patient_id, _parse_date_value


class TestNormalizePatientId:
    def test_none_returns_none(self) -> None:
        assert _normalize_patient_id(None) is None

    def test_nan_returns_none(self) -> None:
        assert _normalize_patient_id(float("nan")) is None

    def test_empty_string_returns_none(self) -> None:
        assert _normalize_patient_id("") is None

    def test_whitespace_only_returns_none(self) -> None:
        assert _normalize_patient_id("   ") is None

    def test_strips_whitespace(self) -> None:
        assert _normalize_patient_id("  P001  ") == "P001"

    def test_plain_string(self) -> None:
        assert _normalize_patient_id("P001") == "P001"

    def test_integer_converts_to_string(self) -> None:
        assert _normalize_patient_id(12345) == "12345"


class TestParseDateValue:
    def test_none_returns_none(self) -> None:
        assert _parse_date_value(None) is None

    def test_nan_returns_none(self) -> None:
        assert _parse_date_value(float("nan")) is None

    def test_empty_string_returns_none(self) -> None:
        assert _parse_date_value("") is None

    @pytest.mark.parametrize(
        "placeholder",
        ["unknown", "Unknown", "unk", "unkown", "n/a", "none", "null"],
    )
    def test_placeholder_strings_return_none(self, placeholder: str) -> None:
        assert _parse_date_value(placeholder) is None

    def test_datetime_object(self) -> None:
        assert _parse_date_value(datetime(2023, 1, 15)) == pd.Timestamp("2023-01-15")

    def test_date_object(self) -> None:
        assert _parse_date_value(date(2023, 1, 15)) == pd.Timestamp("2023-01-15")

    def test_timestamp_object(self) -> None:
        ts = pd.Timestamp("2023-01-15")
        assert _parse_date_value(ts) == ts

    def test_yyyy_mm_dd_string(self) -> None:
        assert _parse_date_value("2023-01-15") == pd.Timestamp("2023-01-15")

    def test_yyyy_dd_mm_string(self) -> None:
        # 15 cannot be a month, so this is unambiguously day-first
        assert _parse_date_value("2023-15-01") == pd.Timestamp("2023-01-15")

    def test_dd_mm_yyyy_string(self) -> None:
        assert _parse_date_value("15-01-2023") == pd.Timestamp("2023-01-15")

    def test_mm_dd_yyyy_string(self) -> None:
        # 15 cannot be a month, so this is unambiguously day-second
        assert _parse_date_value("01-15-2023") == pd.Timestamp("2023-01-15")


class TestGenerateShiftMappings:
    def test_returns_correct_columns(self) -> None:
        result = generate_shift_mappings(["P001"])
        assert list(result.columns) == ["patient_id", "shift_days"]

    def test_one_row_per_patient(self) -> None:
        ids = ["P001", "P002", "P003"]
        result = generate_shift_mappings(ids)
        assert len(result) == 3
        assert list(result["patient_id"]) == ids

    def test_reproducible_with_seed(self) -> None:
        ids = ["P001", "P002", "P003"]
        result1 = generate_shift_mappings(ids, seed=42)
        result2 = generate_shift_mappings(ids, seed=42)
        assert list(result1["shift_days"]) == list(result2["shift_days"])

    def test_different_seeds_differ(self) -> None:
        ids = ["P001", "P002", "P003"]
        result1 = generate_shift_mappings(ids, seed=1)
        result2 = generate_shift_mappings(ids, seed=2)
        assert list(result1["shift_days"]) != list(result2["shift_days"])

    def test_shifts_within_specified_range(self) -> None:
        ids = [f"P{i:03d}" for i in range(100)]
        result = generate_shift_mappings(
            ids, min_shift_days=-7, max_shift_days=7, seed=42
        )
        assert result["shift_days"].between(-7, 7).all()

    def test_empty_patient_list(self) -> None:
        result = generate_shift_mappings([])
        assert len(result) == 0


class TestLoadShiftMappings:
    def test_loads_valid_csv(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "shifts.csv"
        csv_file.write_text("patient_id,shift_days\nP001,5\nP002,-3\n")
        result = load_shift_mappings(str(csv_file))
        assert list(result["patient_id"]) == ["P001", "P002"]
        assert list(result["shift_days"]) == [5, -3]

    def test_raises_on_missing_columns(self, tmp_path: Path) -> None:
        csv_file = tmp_path / "bad.csv"
        csv_file.write_text("id,days\nP001,5\n")
        with pytest.raises(ValueError, match="patient_id.*shift_days"):
            load_shift_mappings(str(csv_file))


class TestGetRowValuesResolvingMerged:
    """Unit tests for _get_row_values_resolving_merged (merged cell header handling)."""

    def test_merged_row_returns_value_repeated_for_each_column(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "crazy"
        ws.merge_cells("A1:D1")
        result = _get_row_values_resolving_merged(ws, 1, 4)
        assert result == ["crazy", "crazy", "crazy", "crazy"]

    def test_normal_row_returns_cell_values(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A2"] = "patient_id"
        ws["B2"] = "measurement"
        ws["C2"] = "date_result"
        ws["D2"] = "type"
        result = _get_row_values_resolving_merged(ws, 2, 4)
        assert result == ["patient_id", "measurement", "date_result", "type"]

    def test_partial_merge_second_column_gets_top_left_value(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "a"
        ws["B1"] = "b"
        ws.merge_cells("B1:C1")
        ws["D1"] = "d"
        result = _get_row_values_resolving_merged(ws, 1, 4)
        assert result == ["a", "b", "b", "d"]


class TestDescriptionMergedRanges:
    """Unit tests for _description_merged_ranges."""

    def test_returns_merged_ranges_within_description_rows(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "title"
        ws.merge_cells("A1:D1")
        result = _description_merged_ranges(ws, 1)
        assert result == ["A1:D1"]

    def test_excludes_merged_ranges_below_description_rows(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "title"
        ws.merge_cells("A1:D1")
        ws["A2"] = "x"
        ws.merge_cells("A2:B2")
        result = _description_merged_ranges(ws, 1)
        assert result == ["A1:D1"]

    def test_empty_when_no_merged_cells(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "a"
        result = _description_merged_ranges(ws, 2)
        assert result == []

    def test_empty_when_num_description_rows_zero(self) -> None:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws["A1"] = "x"
        ws.merge_cells("A1:B1")
        result = _description_merged_ranges(ws, 0)
        assert result == []


class TestShiftExcelDatesWithComplexLayout:
    """Integration tests: header_row, skip_rows_after_header, merged cells, 
    patient sheet from config."""

    def _make_excel_with_complex_header(
        self, path: Path, sheet_name: str = "patients"
    ) -> None:
        """Create an xlsx with: row0 merged title, row1 description, row2 header, row3 
        data-type, row4+ data."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = sheet_name
        ws["A1"] = "crazy"
        ws.merge_cells("A1:D1")
        ws["A2"] = "description"
        ws["B2"] = "description"
        ws["C2"] = "description"
        ws["D2"] = "description"
        ws["A3"] = "patient_id"
        ws["B3"] = "measurement"
        ws["C3"] = "date_result"
        ws["D3"] = "type"
        ws["A4"] = "stupid"
        ws["B4"] = "stupid"
        ws["C4"] = "stupid"
        ws["D4"] = "stupid"
        ws["A5"] = "P001"
        ws["B5"] = "Hb"
        ws["C5"] = "2023-01-15"
        ws["D5"] = "lab"
        ws["A6"] = "P002"
        ws["B6"] = "WBC"
        ws["C6"] = "2023-02-20"
        ws["D6"] = "lab"
        wb.save(path)

    def test_header_row_and_skip_rows_after_header_produce_correct_columns(
        self, tmp_path: Path
    ) -> None:
        """With header_row=2 and skip_rows_after_header=[3], 
        column names come from row 2, row 3 excluded."""
        xlsx = tmp_path / "in.xlsx"
        self._make_excel_with_complex_header(xlsx)
        out = tmp_path / "out.xlsx"
        sheet_configs = {
            "patients": {
                "patient_id_col": "patient_id",
                "date_columns": ["date_result"],
                "header_row": 2,
                "skip_rows_after_header": [3],
            },
        }
        shift_excel_dates(
            input_file=str(xlsx),
            output_file=str(out),
            patient_sheet="patients",
            patient_id_col="patient_id",
            sheet_configs=sheet_configs,
            linking_table_output=str(tmp_path / "linking.csv"),
            seed=42,
        )
        df = pd.read_excel(str(out), sheet_name="patients", header=2)
        assert list(df.columns) == ["patient_id", "measurement", "date_result", "type"]
        # Data should be 2 rows (P001, P002), not 3 (no "stupid" row)
        assert len(df) == 2
        assert list(df["patient_id"]) == ["P001", "P002"]

    def test_patient_sheet_uses_config_header_when_sheet_in_sheet_configs(
        self, tmp_path: Path
    ) -> None:
        """When patient_sheet is in sheet_configs, its header_row is used
         (no patient_header_row arg)."""
        xlsx = tmp_path / "in.xlsx"
        self._make_excel_with_complex_header(xlsx)
        out = tmp_path / "out.xlsx"
        sheet_configs = {
            "patients": {
                "patient_id_col": "patient_id",
                "date_columns": ["date_result"],
                "header_row": 2,
                "skip_rows_after_header": [3],
            },
        }
        """
         Do not pass patient_header_row; patient sheet "patients" 
         is in config with header_row=2
         """
        shift_excel_dates(
            input_file=str(xlsx),
            output_file=str(out),
            patient_sheet="patients",
            patient_id_col="patient_id",
            sheet_configs=sheet_configs,
            linking_table_output=str(tmp_path / "linking.csv"),
            seed=42,
        )
        # Would fail with wrong columns if we had used header_row=0 (row "crazy")
        df = pd.read_excel(str(out), sheet_name="patients", header=2)
        assert "patient_id" in df.columns
        assert "date_result" in df.columns
        assert len(df) == 2


class TestApplyDateShifts:
    def _make_mappings(self, patient_id: str, shift_days: int) -> pd.DataFrame:
        return pd.DataFrame({"patient_id": [patient_id], "shift_days": [shift_days]})

    def test_shifts_date_forward(self) -> None:
        df = pd.DataFrame({"patient_id": ["P001"], "visit_date": ["2023-01-15"]})
        result = apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", 10)
        )
        assert result["visit_date"].iloc[0] == date(2023, 1, 25)

    def test_shifts_date_backward(self) -> None:
        df = pd.DataFrame({"patient_id": ["P001"], "visit_date": ["2023-06-01"]})
        result = apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", -5)
        )
        assert result["visit_date"].iloc[0] == date(2023, 5, 27)

    def test_zero_shift_leaves_date_unchanged(self) -> None:
        df = pd.DataFrame({"patient_id": ["P001"], "visit_date": ["2023-01-15"]})
        result = apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", 0)
        )
        assert result["visit_date"].iloc[0] == date(2023, 1, 15)

    def test_unknown_patient_leaves_date_unchanged(self) -> None:
        df = pd.DataFrame({"patient_id": ["P999"], "visit_date": ["2023-01-15"]})
        result = apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", 10)
        )
        assert result["visit_date"].iloc[0] == date(2023, 1, 15)

    def test_placeholder_date_becomes_none(self) -> None:
        df = pd.DataFrame({"patient_id": ["P001"], "visit_date": ["Unknown"]})
        result = apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", 10)
        )
        assert result["visit_date"].iloc[0] is None

    def test_none_date_stays_none(self) -> None:
        df = pd.DataFrame({"patient_id": ["P001"], "visit_date": [None]})
        result = apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", 10)
        )
        assert result["visit_date"].iloc[0] is None

    def test_missing_date_column_is_skipped(self) -> None:
        df = pd.DataFrame({"patient_id": ["P001"], "visit_date": ["2023-01-15"]})
        mappings = self._make_mappings("P001", 10)
        # "nonexistent" should be silently skipped; "visit_date" still shifted
        result = apply_date_shifts(
            df, "patient_id", ["nonexistent", "visit_date"], mappings
        )
        assert result["visit_date"].iloc[0] == date(2023, 1, 25)

    def test_strips_whitespace_from_patient_ids(self) -> None:
        df = pd.DataFrame({"patient_id": ["  P001  "], "visit_date": ["2023-01-15"]})
        result = apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", 10)
        )
        assert result["visit_date"].iloc[0] == date(2023, 1, 25)

    def test_multiple_patients_shifted_independently(self) -> None:
        df = pd.DataFrame(
            {
                "patient_id": ["P001", "P002"],
                "visit_date": ["2023-01-01", "2023-01-01"],
            }
        )
        mappings = pd.DataFrame({"patient_id": ["P001", "P002"], "shift_days": [5, -5]})
        result = apply_date_shifts(df, "patient_id", ["visit_date"], mappings)
        assert result["visit_date"].iloc[0] == date(2023, 1, 6)
        assert result["visit_date"].iloc[1] == date(2022, 12, 27)

    def test_result_dates_are_date_not_datetime(self) -> None:
        df = pd.DataFrame({"patient_id": ["P001"], "visit_date": ["2023-01-15"]})
        result = apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", 0)
        )
        val = result["visit_date"].iloc[0]
        assert isinstance(val, date)
        assert not isinstance(val, datetime)

    def test_does_not_mutate_input(self) -> None:
        df = pd.DataFrame({"patient_id": ["P001"], "visit_date": ["2023-01-15"]})
        original_visit = df["visit_date"].iloc[0]
        apply_date_shifts(
            df, "patient_id", ["visit_date"], self._make_mappings("P001", 10)
        )
        assert df["visit_date"].iloc[0] == original_visit
