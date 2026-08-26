from pathlib import Path

import pytest
from openpyxl import load_workbook

from nuh_helper import shift_excel_dates_inplace
from nuh_helper.date_shift import RowMissingID


def test_missing_ids(tmp_path: Path) -> None:

    source_file = Path(__file__).parent / "data/blanks/missing_ids.xlsx"
    output_file = tmp_path / "output_file.xlsx"
    linking_table_old = tmp_path / "linking_table_old.csv"
    linking_table_out = tmp_path / "linking_table_out.csv"

    sheet_configs = {
        "sheet": {
            "patient_id_col": "pitd",
            "date_columns": ["dob"],
            "header_row": 0,
            "skip_rows_after_header": [],
        }
    }

    with pytest.raises(RowMissingID) as info:
        shift_excel_dates_inplace(
            input_file=str(source_file),
            output_file=str(output_file),
            patient_sheet="sheet",
            patient_id_col="pitd",
            sheet_configs=sheet_configs,
            min_shift_days=-20,
            max_shift_days=-1,
            seed=14333,
            linking_table_path=str(linking_table_old),
            linking_table_output=str(linking_table_out),
        )

    assert (
        info.value._message
        == "missing id for [sheet_name='sheet', 4, 2] = value='2009-12-05 00:00:00'"
    )


def test_valid(tmp_path: Path) -> None:

    source_file = Path(__file__).parent / "data/blanks/valid.xlsx"
    output_file = tmp_path / "output_file.xlsx"
    linking_table_old = Path(__file__).parent / "data/blanks/valid.csv"
    linking_table_out = tmp_path / "linking_table_out.csv"

    sheet_configs = {
        "sheet": {
            "patient_id_col": "pitd",
            "date_columns": ["dob"],
            "header_row": 0,
            "skip_rows_after_header": [],
        }
    }

    shift_excel_dates_inplace(
        input_file=str(source_file),
        output_file=str(output_file),
        patient_sheet="sheet",
        patient_id_col="pitd",
        sheet_configs=sheet_configs,
        min_shift_days=-20,
        max_shift_days=-1,
        seed=14333,
        linking_table_path=str(linking_table_old),
        linking_table_output=str(linking_table_out),
    )

    sheet = load_workbook(output_file).worksheets[0]

    # first column
    assert sheet.cell(1, 1).value == "pitd"
    assert sheet.cell(2, 1).value is None
    assert sheet.cell(3, 1).value is None
    assert sheet.cell(4, 1).value is None
    assert sheet.cell(5, 1).value == "nuh1"
    assert sheet.cell(6, 1).value == "nuh2"
    assert sheet.cell(7, 1).value is None
    assert sheet.cell(8, 1).value == "nuh3"
    assert sheet.cell(9, 1).value == "nuh4"

    # second column
    assert sheet.cell(1, 2).value == "dob"
    assert sheet.cell(2, 2).value is None
    assert sheet.cell(3, 2).value is None
    assert sheet.cell(4, 2).value is None
    assert str(sheet.cell(5, 2).value) == "2019-08-31 00:00:00"
    assert str(sheet.cell(6, 2).value) == "2001-09-01 00:00:00"
    assert sheet.cell(7, 2).value is None
    assert str(sheet.cell(8, 2).value) == "2001-09-24 00:00:00"
    assert str(sheet.cell(9, 2).value) == "2006-10-28 00:00:00"

    # third column
    assert sheet.cell(1, 3).value == "postcode"
    assert sheet.cell(2, 3).value is None
    assert sheet.cell(3, 3).value is None
    assert sheet.cell(4, 3).value is None
    assert sheet.cell(5, 3).value == "G12"
    assert sheet.cell(6, 3).value == "NG7"
    assert sheet.cell(7, 3).value is None
    assert sheet.cell(8, 3).value == "LH1"
    assert sheet.cell(9, 3).value == "M1"
