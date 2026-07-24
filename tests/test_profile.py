"""Unit tests for nuh_helper.profile"""

from pathlib import Path

from openpyxl import load_workbook

from nuh_helper.profile import FieldTypeInfo, generate_scan_report


def _field_type(values: list[str]) -> str:
    info = FieldTypeInfo()
    for value in values:
        info.add(value)
    return info.type


class TestFieldTypeInfo:
    def test_all_integers_is_int(self) -> None:
        assert _field_type(["1", "22", "-3", "045"]) == "INT"

    def test_mixed_integers_and_decimals_is_real(self) -> None:
        assert _field_type(["1", "2.5", "-3"]) == "REAL"

    def test_scientific_notation_is_real(self) -> None:
        assert _field_type(["1e10", "2.5E-3"]) == "REAL"

    def test_dates_are_detected(self) -> None:
        assert _field_type(["2020-01-15", "2020-02-16", "2020-03-17"]) == "DATE"

    def test_slash_separated_dates_are_detected(self) -> None:
        assert _field_type(["15/01/2020", "16/02/2020"]) == "DATE"

    def test_alphanumeric_is_varchar(self) -> None:
        assert _field_type(["P001", "P002", "abc"]) == "VARCHAR"

    def test_all_empty_is_empty(self) -> None:
        assert _field_type(["", "", ""]) == "EMPTY"

    def test_blank_values_do_not_affect_type(self) -> None:
        assert _field_type(["1", "", "2"]) == "INT"

    def test_long_text_is_text(self) -> None:
        long_value = "a" * 150
        assert _field_type([long_value, long_value]) == "TEXT"

    def test_max_length_tracks_raw_value_length(self) -> None:
        info = FieldTypeInfo()
        info.add("short")
        info.add("a much longer value")
        assert info.max_length == len("a much longer value")


class TestGenerateScanReport:
    def test_infers_column_types(self, tmp_path: Path) -> None:
        csv_path = tmp_path / "patients.csv"
        csv_path.write_text(
            "patient_id,age,measurement,dob,name\n"
            "P001,45,120,2020-01-15,Alice\n"
            "P002,50,130,2020-02-16,Bob\n"
            "P003,60,110,2020-03-17,Carol\n"
        )

        output_path = generate_scan_report(
            [str(csv_path)], output_path=str(tmp_path / "ScanReport.xlsx")
        )

        wb = load_workbook(output_path)
        rows = list(wb["Field Overview"].iter_rows(values_only=True))
        types = {row[1]: row[3] for row in rows[1:] if row[1] is not None}

        assert types == {
            "patient_id": "VARCHAR",
            "age": "INT",
            "measurement": "INT",
            "dob": "DATE",
            "name": "VARCHAR",
        }
