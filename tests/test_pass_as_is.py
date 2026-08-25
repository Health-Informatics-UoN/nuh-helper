from pathlib import Path

import pytest
from openpyxl import load_workbook

from nuh_helper import shift_excel_dates_inplace
from nuh_helper.date_shift import ShiftFoundNonDate


@pytest.mark.parametrize("allow_passthrough", [True, False])
def test_just_shift(allow_passthrough: bool, tmp_path: Path) -> None:
    sheet_configs = {
        "page-data": {
            "patient_id_col": "pid",
            "date_columns": ["dob"],
            "header_row": 1,
            "skip_rows_after_header": [],
        }
    }

    output_path = tmp_path / "target.xlsx"

    def body() -> None:
        shift_excel_dates_inplace(
            input_file=str(Path(__file__).parent / "data/passed/workbook.xlsx"),
            output_file=str(output_path),
            patient_sheet="page-data",
            patient_id_col="pid",
            sheet_configs=sheet_configs,
            min_shift_days=-20,
            max_shift_days=-1,
            seed=14333,
            linking_table_path=str(Path(__file__).parent / "data/passed/offsets.csv"),
            linking_table_output=str(tmp_path / "offsets.csv"),
        )

    if not allow_passthrough:
        with pytest.raises(ShiftFoundNonDate) as info:
            body()
        assert str(info.value) == "page='page-data'[5, 2 @ col_name='dob'] val='mssing'"
        return
    else:
        sheet_configs["page-data"]["pass_as_is"] = {"dob": ["mssing"]}
        body()

    workbook = load_workbook(output_path)

    worksheet = workbook.worksheets[1]

    # first column
    assert worksheet.cell(1, 1).value == "personal id"
    assert worksheet.cell(2, 1).value == "pid"
    assert worksheet.cell(3, 1).value == "nuh71"
    assert worksheet.cell(4, 1).value == "nuh06"
    assert worksheet.cell(5, 1).value == "nuh67"
    assert worksheet.cell(6, 1).value == "nuh27"

    # last column
    assert worksheet.cell(1, 3).value == "pizza topping"
    assert worksheet.cell(2, 3).value == "top"
    assert worksheet.cell(3, 3).value == "cheese"
    assert worksheet.cell(4, 3).value == "unknown"
    assert str(worksheet.cell(5, 3).value) == "2016-09-17 00:00:00"
    assert worksheet.cell(6, 3).value == "2016-07-18 idk"

    # the important column to check - the dates
    assert worksheet.cell(1, 2).value == "birthday"
    assert worksheet.cell(2, 2).value == "dob"

    assert str(worksheet.cell(3, 2).value) == "2001-12-17 00:00:00"
    assert str(worksheet.cell(4, 2).value) == "1993-09-20 00:00:00"
    assert str(worksheet.cell(6, 2).value) == "1999-11-30 00:00:00"

    # change that's under test
    assert worksheet.cell(5, 2).value == "mssing"
